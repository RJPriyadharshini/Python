"""import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
driver.implicitly_wait(10)

# Amount of time driver should wait to find the element(if you call next condition it perform)
# implicitly_wait - it is applicable to all the below statements


driver.get("https://opensource-demo.orangehrmlive.com/")

driver.find_element(By.NAME, "username").send_keys("Admin")
driver.find_element(By.NAME,"password").send_keys("admin123")
driver.find_element(By.XPATH, "//button[@type='submit']").click()

#XPATH SYNTAX -     //tagname[@Attribute='Value']

actual_title = driver.title
expected_title = "OrangeHRM"

if actual_title == expected_title:
    print("Login Test Passed")
else:
    print("Login Test Failed")


time.sleep(5)

# performance of the script is very poor
# if the element is not available within the time mentioned


driver.close()  """

"""DAY -2


LOCATORS 

Identify elements in the webpage we use "locators"

Locators 

  1.id 
  2.Name
  3.Link text
  4.Partial Link text
  5.ClassName
  6.TagName
  
Customized Locators - CSS SELECTOR AND XPATH


CSS SELECTOR 

1.Tag and ID                 - tagname#valueofID
2.Tag and class              - tagname.valueofclass
3.Tag and attribute          - tagname[attribute=value]
4.Tag, class and attribute   - tagname[valueofclass[attribute=value]]


XPATH 

1.Absolute XPath(Full)
2.Relative XPath(Partial)

close()  - close only one browser at a time
quit() - close all browsers

HTML STRUCTURE

<input name = "txtUsername" id="txtUsername" type="text"(input is element , name is attribute , txtusername is value
 """

"""
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://demo.nopcommerce.com/")
driver.implicitly_wait(20)
driver.maximize_window()
driver.find_element(By.ID,"small-searchterms").send_keys("Lenovo Thinkpad X1 Carbon Laptop")
driver.find_element(By.XPATH,"//button[@type='submit']").click()
driver.find_element(By.LINK_TEXT,"Register").click()
#driver.find_element(By.PARTIAL_LINK_TEXT,"Reg").click()   [both are same we can use link or partial link]
#driver.find_element(By.XPATH,"//a[@lehref='/novo-thinkpad-x1-carbon-laptop']").click()
time.sleep(50)
driver.close()"""

#find elements - find all the elements with same tag  - multiple web elements

"""sliders=driver.find_elements(By.CLASS_NAME,'homeslider-container')
print(len(slider))   -> print the total len of the slider 

links=driver.find_elements(By.TAG_NAME,'a')
print(len(links))   total num of links present """

#CSS SELECTORS (tag is optional)

"""import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://www.facebook.com/")
driver.implicitly_wait(20)

#TAG AND ID - CSS  [tagname#valueofId]

#driver.find_element(By.CSS_SELECTOR,"input#email").send_keys("aba")
#driver.find_element(By.CSS_SELECTOR,"#email").send_keys("tag and id")

#TAG AND CLASS  [tagname.valueofclass]

#driver.find_element(By.CSS_SELECTOR,"input.inputtext").send_keys("Tag and class")
#(actual class value is "inputtext _55r1 _6luy" sometimes we have to remove the text after the space)

#Tag and attribute [tagname[attribute=value]    ]

#driver.find_element(By.CSS_SELECTOR,"input[data-testid=royal_email]").send_keys("tag attributes")

#Tag class and attribute  [ tagname.valueofclass[attribute=value]   ]

driver.find_element(By.CSS_SELECTOR,"input.inputtext[data-testid=royal_pass").send_keys("pass")
#(we have both same tagname and class in that time we have to identify diff attributed for that two selection)

time.sleep(100)
driver.close()   """

#  DAY -3

"""XPATH (works based on DOM)

-> Finding an element in webpage
-> finding element using HTML DOM Structure
-> navigate through elements and attributes
-> address of the element

DOM-Document object model

->API interface it represents the structure of  HTML and XML Documents as tree structure
->used to locate the web elemets,and creats a DOM of the page


Absolute path (copy full xpath)  

-> use only tags / nodes
-> Start from root node

eg:

/html/body/nav/div/div[2]/div[2]/ul/li[2]/a/button

Relative path  

-> directly to the element
-> use attributes

eg:

//*[@id="navbarSupportedContent"]/div[2]/ul/li[2]/a/button

Write XPATH Manually:

Absolute xpath
/html/body/div[6]/div[1]/div[1]/div[2]/div[1]/ul/li[1]/a

relative
 //tagname[@Attribute='Value']
 
 
selectors hub - extension used in chrome,browser (to identify tags)

Interview question:

Most of the time we use relative XPath because ?

-> If developer introduced the new element then absolute xpath will be broken
-> if developer changed the location then absolute xpath will be broken
-> absolute xpath is unstable

XPATH OPTIONS

1. and 
2. or
3. contains
4. starts-with
5. text

if the button is "START/STOP" how to identify at that time we use contains or startswith
//*[contains(@id,st)] or //*[starts-with(@id,'st')]   or //*[@id='start' or id='stop'] - in this if you call it start and stop both will execute



import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC



driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://www.orangehrm.com/")
# relative xpath
driver.find_element(By.XPATH, "//button[@class='btn btn-ohrm btn-free-demo']").click()

driver.quit()
time.sleep(100)  """

# DAY - 4

"""XPATH AXES  - travels from top to bottom and bottom to top in the page

self     
     - 
-> whichever node we start from
-> SYNTAX - //*[attribute=’value’]/self::tagname

parent

-> traverse parent element of the current html tag
-> //*[attribute=’value’]/parent::parent::tagname

child

-> traverse all child elements of the current html tag.
-> SYNTAX - //*[attribute=’value’]/child::tagname

ancestor

-> Traverse all the ancestor elements(grandparent,parent etc ) of the current html tag
-> SYNTAX - //*[attribute=’value’]/ancestor::tagname

descendant

-> Traverse all the ancestor elements(child, grandchild etc ) of the current html tag
-> SYNTAX - //*[attribute=’value’]/descendant::tagname

following

-> traverse all element that comes after the  current html tag.
-> SYNTAX - //*[attribute=’value’]/following::tagname

following-sibling

-> Traverse from current html tag to next sibling html tag
-> //current html tag[@attribute=’value’]/following-sibling::sibling tag [@attribute=’value’]

preceding

-> traverse all nodes that comes before the current html tag.
-> SYNTAX - //*[attribute=’value’]/preceding::tagname

preceding-sibling

-> Traverse from current html tag to previous sibling html tag
-> //current html tag[@attribute=’value’]/preceding-sibling::previous tag [@attribute=’value’]



import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://money.rediff.com/gainers/bse/daily/groupa")
driver.implicitly_wait(20)

# self

#text_msg=driver.find_element(By.XPATH, "//a[normalize-space()='Angel One']/self::a").text
#print(text_msg)

#parent

#p_msg=driver.find_element(By.XPATH, "//a[normalize-space()='Angel One']/parent::td").text
#print(p_msg)

#child

# our self element not have child so we go to ancestor and pick that child element

child=driver.find_elements(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr/child::td")
print(len(child)) # return all the elements to that td category

#ancestor

# child element data will show
ar=driver.find_element(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr")
print(ar)

#descendant

des=driver.find_elements(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr/descendant::*")
print(len(des))  #td-5 is descendnt of tr

#following(below ele)

fol=driver.find_elements(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr/following::*")
print(len(fol))

#following sibling   (subset of following)

#we dont have siblings for self so we go the ancestor and find siblings

fol_sib=driver.find_elements(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr/following-sibling::*")
print(len(fol_sib))

#preceding(above ele)

pre=driver.find_elements(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr/preceding::tr")
print(len(pre))

#preceding siblings

pre_sib=driver.find_elements(By.XPATH, "//a[normalize-space()='Angel One']/ancestor::tr/preceding-sibling::tr")
print(len(pre_sib))

driver.quit()
time.sleep(100)
"""


# DAY - 5


"""
1. Application commands               - Get , title , current_URL , Page_source
2. Conditional commands               - is_enabled,is_disabled,is_selected
3. Browser commands                   - Close , quit
4. Navigational commands              - Bach , refresh , forward
5. Wait commands                      - implicit and explicit


Application commands:

1. get         - opening the app URL
2. title       - to capture the title of current webpage
3. current_url - capture current_url of the web page
4. page_source - capture source code of the page

application elements done by driver

Conditional commands 

1. is_displayed   - element is present in the webpage or not   
2. is_enabled     - element is enabled(element do some action) returns true.
3. is_selected    - check radio and checkboxes is selected returns true or false

conditional elements done by web_elements



BROWER COMMANDS

Close()  -  simply close the browser but backend process is running , close one browser at a time
Quit()  -  close browser along with process . close all the browsers

Navigational commands

1. back
2. forward
3. refresh


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://opensource-demo.orangehrmlive.com/")
print(driver.title)
print(driver.current_url)
print(driver.page_source)


driver.quit()   """


"""driver.get("https://demo.nopcommerce.com/register")
search=driver.find_element(By.XPATH,"//input[@id='small-searchterms']")
print("Display status:",search.is_displayed())       # True
print("enabled status:",search.is_enabled())         # True

#is_selected for radio and checkbox

rd_female=driver.find_element(By.XPATH,"//input[@id='gender-female']")
rd_male=driver.find_element(By.XPATH,"//input[@id='gender-male']")

rd_female.click()
print(rd_female.is_selected())        # True
print(rd_male.is_selected())          # False

time.sleep(100)
driver.quit()


#navigational commands

driver.get("https://opensource-demo.orangehrmlive.com/")
driver.get("https://www.amazom.com")

driver.back()      #back to orange app
driver.forward()   #go to amazon
driver.refresh()   #refresh the page
time.sleep(10)
driver.quit()  


driver.get ("https://demo.nopcommerce.com/")
# find element -- matching with one element

element_var=driver.find_element(By.XPATH,"//input[@id='small-searchterms']")
element_var.send_keys("Apple Phone")

# locators matching with the multiple web elements

element_var=driver.find_element(By.XPATH,"//div[@class='footer']//a")
print(element.text)   # only print one web element because of find_element

#element not available them throw NosuchElementException
#find element is not able to find the element at that time it shows NosuchElementException

login=driver.find_element(By.LINK_TEXT,"Log")  #Log in is correct
login.click()



time.sleep(10)
driver.quit()   

######## find_elements() ##### - Return multiple webelements

# Locators match with single webelement

#elements=driver.find_elements(By.XPATH,"//input[@id='small-searchterms']")
#print(len(elements)) #list collection we cannot able to access elements
#print(elements[0].send_keys())  # we extract web elements and send actions

# locators matching with multiple web elements

driver.get("https://demo.nopcommerce.com/")
element_var=driver.find_elements(By.XPATH,"//div[@class='footer']//a")
print(len(element_var))

for ele in element_var:
    print(ele.text)


#element not avaliable

ele=driver.find_elements(By.LINK_TEXT,"Log")
print(len(ele))  # not show any exception

time.sleep(10)
driver.quit() 

######text vs get_attribute(values)#####

driver.get("https://demo.nopcommerce.com/")
email=driver.find_element(By.XPATH,"//input[@id-'Email']")
email.clear()  # default email is cleared
email.send_keys("abc@gmail.com")
print("result of text",email.text)   # Print nothing
print("result of get_attribute",email.get_attribute('value'))

#<input id="123"  name="xyz" > Email:</input>  .... Email is inner text
# Text - returns Inner text of the element
# get_attribute - returns values of any attribute of web element

log=driver.find_element(By.XPATH,"//button[normalize-space()='Log in']")
print("result of text",log.text)
print("result of get_attribute",log.get_attribute('value'))
print("result of get_attribute",log.get_attribute('type')) """


####### DAY -6 ########

# Wait commands - used for syrnchronizing prblm

"""
time.sleep (maximum time wait for that element)

advantage 
-> simple to use

Disadvantage

-> performance of the script is very poor
-> if the element is not available within the time mentioned still there is a chance of getting exception 

1.implicit wait(wait for element if element is available proceed to next)

ADVANTAGE

-> Single statement
-> performance will not be reduced (if the element is available within the time it proceed to execute further
DISADVANTAGE
-> If the element is not available within the time mentioned , still there is a chance of getting exception

# Amount of time driver should wait to find the element(if you call next condition it perform
# implicitly_wait - - it is applicable to all the below statements


2.explicit wait

explicit waits - it works based on condition
mywait=WebDriverWait(driver,10,ignored_exceptions=[NoSuchElementException,ElementNotVisibleException,ElementsNotSelectableException,Exception])
poll_frequency=2  - total 10 sec in tha every 2 sec it goes and try to find that element it reduce time
mywait=WebDriverWait(driver,10,poll_frequency,ignored_exceptions=[NoSuchElementException,ElementNotVisibleException,ElementsNotSelectableException,Exception])

Adv
-> More effectively work

Dis 
-> Multiple places enter feels some difficulty


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome()
mywait=WebDriverWait(driver,10)  # explicit wait declaration

#driver.implicitly_wait(20)
driver.get("https://www.google.com/")
searchbox=driver.find_element(By.NAME,'q')
searchbox.send_keys("Selenium")
#searchbox.submit()     # Press enter key

#driver.find_element(By.XPATH,"//span[normalize-space()='selenium']").click()
#driver.find_element(By.XPATH,"//h3[text()='Selenium']").click()
mywait.until(EC.presence_of_element_located((By.XPATH,"//h3[text()='Selenium']")))
mywait.click()# condition is stastified it will execute
#If the condition is not true until it wait for that element
time.sleep(10)
driver.quit() """


####DAY - 7 #######

"""import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

#Select checkboxes

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://itera-qa.azurewebsites.net/home/automation")
#driver.find_element(By.XPATH,"//input[@id='monday']").click()


checkboxes=driver.find_elements(By.XPATH,"//input[@type='checkbox' and contains(@Id,'day')]")
print(len(checkboxes))


#for i in range(len(checkboxes)):
    #checkboxes[i].click()

#for checkbox in checkboxes:
    #checkbox.click()

# Select multiple

for checkbox in checkboxes:
    weekname=checkbox.get_attribute('id')
    if weekname=="monday" or weekname=="sunday":
        checkbox.click()

#Select two checkboxes from last

for i in range(len(checkboxes)-2,len(checkboxes)):  #Total 7 [range(5,7)]
    checkboxes[i].click()   # sat and sun

time.sleep(10)
driver.quit()


# select first 2 checkboxes

for i in range(2):
    checkboxes[i].click()
#or

for i in range(len(checkboxes)):
    if i<2:
        checkboxes[i].click()

#for select and unselect using the same click fun


#if selected remove that if unselected add that

for checkbox in checkboxes:
    if checkbox.is_selected():
        checkbox.click()
    else:
        checkbox.click()



######   LINKS   ###

1.Internal link   - Link available on same page(naviagte to same page)
2.External link   - link navigate to other webpage
3.Broken link     - Target link is not available


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://demo.nopcommerce.com/")
driver.implicitly_wait(20)
driver.find_element(By.LINK_TEXT,"Digital downloads").click()
#driver.find_element(By.PARTIAL_LINK_TEXT,"Digital").click()

links=driver.find_elements(By.TAG_NAME,'a')
print(len(links))

for link in links:
    print(link.text)  # link is web element we cannot directly print so that we use text
time.sleep(100)
driver.quit()


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import requests as request

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("http://www.deadlinkcity.com/")

all_links=driver. find_elements(By.TAG_NAME,'a')
count=0

for link in all_links:
    url=link.get_attribute('href')
    try:
        res = request.get(url)  # send a request to the server
    except:
        None
    if res.status_code >= 400 and res.status_code < 406:
        print(url," is broken link")
        count+=1
    else:
        print(url, "is not broken")

print("total is:",count)

time.sleep(100)
driver.quit()

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://www.opencart.com/index.php?route=account/register")

# Find the dropdown element

drp_coun_els = driver.find_element(By.XPATH, "//select[@id='input-country']").click()
drp_coun = Select(drp_coun_els)

#select option from dropdown

1.select_by_visible_text
2.select_by_index
3.select_by_value


#drp_coun.select_by_visible_text("Angola")
#drp_coun.select_by_value("6")
drp_coun.select_by_index(6) #count by manual



all_opt=drp_coun.options
print("all options is:",len(all_opt))
print(all_opt)

# Iterate through options and print their text

for option in all_opt:
    print(option.text)

#select option from drop down without using build-in functions

for option in all_opt:
    if option.text=="Angola":
        option.click()
        break

time.sleep(200)
driver.quit()


####### DAY - 8 #####

#   Alerts/Popups

alert = driver.switch_to.alert
alert.text
alert.accept
alert.dismiss


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://the-internet.herokuapp.com/javascript_alerts")
# open alert window
driver.find_element(By.XPATH,"//button[normalize-space()='Click for JS Prompt']").click()
time.sleep(5)
alt_win=driver.switch_to.alert  # switch to alert window
alt_win.text
alt_win.send_keys("User input 'Welcome'")
alt_win.accept() #close alt window by using ok button.
#alt_win.dismiss() #close alt window using cancel button

time.sleep(50)
driver.quit()

#########

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
#driver.implicitly_wait(20)
driver.get("https://mypage.rediff.com/login/dologin")
driver.find_element(By.XPATH,"//input[@value='Login']").click()
time.sleep(5)
#alt_win=driver.switch_to.alert
#alt_win.accept()
#or

driver.switch_to.alert.accept()
time.sleep(50)
driver.quit()



# Authentication popup

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
#driver.implicitly_wait(20)
driver.get("https://theinternet.herokuapp.com/basic_auth")  #normal url
#driver.get("https://admin:admin@theinternet.herokuapp.com/basic_auth")   # value passed in that URL
time.sleep(50)
driver.quit() 

# Handle frames or Iframes

#SWITCH TO FRAMES

switch_to.frame(name of the frame)
switch_to.frame(id of the frame)
switch_to.frame(web element)
switch_to.frame(0) by using index 

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://www.selenium.dev/selenium/docs/api/java/index.html?overview-summary.html")
driver.switch_to.frame("packageListFrame")
driver.find_element(By.LINK_TEXT,"org.openqa.selenium").click()
driver.switch_to.default_content()   #go back to main page [driver cannot move one frame to another frame]
driver.switch_to.frame("packageFrame")
driver.find_element(By.LINK_TEXT,"WebDriver").click()
driver.switch_to.default_content() 
driver.switch_to.frame("classFrame")
driver.find_element(By.XPATH,"//li[@class='nav-bar-cell1-rev']").click()

time.sleep(50)
driver.quit()

###########

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
driver.get("https://demo.automationtesting.in/Frames.html")
driver.implicitly_wait(20)
driver.find_element(By.XPATH,"//a[normalize-space()='Iframe with in an Iframe']").click()

outerframe=driver.find_element(By.XPATH,"//iframe[@src='MultipleFrames.html']")
driver.switch_to.frame(outerframe)

innerframe=driver.find_element(By.XPATH,"/html/body/section/div/div/iframe")
driver.switch_to.frame(innerframe)

driver.find_element(By.XPATH,"//input[@type='text']").send_keys("Welcome")
driver.switch_to.frame() # have one parent frame(outer i frame)

time.sleep(50)
driver.quit()

#### BROWSE WINDOWS ####

switch to window

current_window_handle - window id of single browser window
window_handles - window id of multiple browser windows 

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://opensource-demo.orangehrmlive.com/")
driver.maximize_window()

#windowid=driver.current_window_handle
#print(windowid)  #454CD7756633724EF896F93D4BDAA09E
# every time it gives a Diff windowid

driver.find_element(By.LINK_TEXT,"OrangeHRM, Inc").click()
windowids=driver.window_handles


## APPROACH 1

parentwindowid=windowids[0]
childwindowid=windowids[1]

print(parentwindowid,childwindowid)

driver.switch_to.window((childwindowid))
print("title of child: ",driver.title)

driver.switch_to.window(parentwindowid)
print("title of parent: ",driver.title)

## APPROACH 2

for win in windowids:
    driver.switch_to.window(win)
    print(driver.title)
    if driver.title=="OrangeHRM":
        driver.close()    # close the child window

time.sleep(20)
"""

##### DAY -9 ###

"""import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

chrome_options = Options()
chrome_options.add_argument("--disable-notifications")
driver = webdriver.Chrome()
driver.get("https://whatmylocation.com/")

driver.quit()


###WEB TABLE####
 1. Static web table
 2. Dynamic webtable
 td - table data
 
 

import time
from selenium import webdriver
from selenium.webdriver.common.by import By

# Create a Chrome WebDriver instance
driver = webdriver.Chrome()

# Navigate to the website
driver.get("https://testautomationpractice.blogspot.com/")

# Find the number of rows and columns in the table
rows = driver.find_elements(By.XPATH, "//table[@name='BookTable']//tr")
columns = driver.find_elements(By.XPATH, "//table[@name='BookTable']//tr[1]/th")
num_rows = len(rows)
num_columns = len(columns)

print("Number of rows:", num_rows)
print("Number of columns:", num_columns)

# READ SPECIFIC ELEMENT
specific = driver.find_element(By.XPATH, "//table[@name='BookTable']/tbody/tr[5]/td[1]").text
print("Specific element:", specific)

# Read all the elements

for i in range(2, num_rows + 1):          #2,3,4,5,6,7,8
    for j in range(1, num_columns + 1):    # 1,2   1,3  1,4
        specific = driver.find_element(By.XPATH, "//table[@name='BookTable']/tbody/tr["+str(i)+"]/td["+str(j)+"]").text         #1st convert to str , enter values in double quotes , add + before and after the value
        print(specific)
    print()

# Read data based on condition

for i in range(2, num_rows + 1):          #2,3,4,5,6,7,8
    Author=specific = driver.find_element(By.XPATH, "//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[2]").text
    if Author=="Amit":
        a=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[2]/td[1]").text
        b=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[7]/td[1]").text
        print("author is:",Author,"  ","book is:",a)
        print("author is:",Author,"  ","book is:",b)
    elif Author=="Mukesh":
        book=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[3]/td[1]").text
        price=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[3]/td[4]").text
        print("book is:",book, "    ","price is:",price)


driver.quit()


## DAY -10 ####

DATE PICKER
1. Standard
2. Non-standard (Customized)

#mm/dd/year

import time
from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://jqueryui.com/datepicker/")
driver.switch_to.frame(0)
driver.find_element(By.XPATH,"//input[@id='datepicker']").click()

date="11"
month="March"
year="2020"

while True:
    mon=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-month']").text
    yr=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-year']").text

    if mon==month and yr==year:
        break
    else:
        driver.find_element(By.XPATH,"//*[@id='ui-datepicker-div']/div/a[1]/span").click()


dates=driver.find_elements(By.XPATH," //table[@class='ui-datepicker-calendar']/tbody/tr/td/a")
for ele in dates:
    if ele.text==date:
        ele.click()
        break

time.sleep(20)
driver.quit()


###### DAY -11 ####

Mouse operations

1. ActionChains   -  move_to_element(element)
2. Right click    -  context_click(element)
3. Double click   -  double_click(element)
4. Drag and drop  -  drag_and_drop(source,target)

drag_and_drop_by_offset   - for range 

import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://opensource-demo.orangehrmlive.com/")
driver.find_element(By.NAME, "username").send_keys("Admin")
driver.find_element(By.NAME, 'password').send_keys("admin123")
driver.find_element(By.XPATH, "//button[@type='submit']").click()

admin=driver.find_element(By.XPATH,"//*[@id='menu_admin_viewAdminModule']/b")
user_man=driver.find_element(By.XPATH,"//*[@id='menu_admin_UserManagement']")
users=driver.find_element(By.XPATH,"//*[@id='menu_admin_viewSysytemUsers']")

#MOUSEHOVER

act=ActionChains(driver)   # pass driver
#mouseHover
# Perform is one method by that only action will execute
act.move_to_element(admin).move_to_element(user_man).move_to_element(users).click().perform()
driver.quit()


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains


driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://swisnl.github.io/jQuery-contextMenu/demo.html")
Button=driver.find_element(By.XPATH,"//span[normalize-space()='right click me']")
#right click
act=ActionChains(driver)
act.context_click(Button).perform()
driver.find_element(By.XPATH,"/html/body/ul/li[1]").click()
alert=driver.switch_to.alert
alert.accept()
driver.implicitly_wait(2)

time.sleep(10)
driver.quit()

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://www.w3schools.com/tags/tryit.asp?filename=tryhtml5_ev_ondblclick3")
driver.switch_to.frame("iframeResult")   # switch to frame

field1 = driver.find_element(By.XPATH, "//input[@id='field1']")
field1.clear()
field1.send_keys("Welcome")

# Corrected XPath for the button
button = driver.find_element(By.XPATH, "//button[text()='Copy Text']")

act = ActionChains(driver)
act.double_click(button).perform()

field2=driver.find_element(By.XPATH,"//input[@id='field2']")

if field1.text==field2.text:
    print("matched")
else:
    print("unmatched")
    
time.sleep(2)  # Add a small delay to see the changes before quitting the browser

driver.quit()



import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://demoqa.com/droppable/")
sour_ele=driver.find_element(By.XPATH,"//*[@id='draggable']")
tar_ele=driver.find_element(By.XPATH,"//*[@id='droppable']")
# drag and drop
act=ActionChains(driver)
act.drag_and_drop(sour_ele,tar_ele).perform()
driver.close()


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://www.jqueryscript.net/demo/Price-Range-Slider-jQuery-UI/")

# Range drag min to max( 1 to 500)

min=driver.find_element(By.XPATH,"//body//div//span[1]")
max=driver.find_element(By.XPATH,"//body//div//span[2]")

print("Before Moving.....")
print(min.location)  # X and Y axis  {x:59 , y:252}
print(max.location)  # (x:412, y:250)

# what we added or reduced with x, it is an approx value nearby range will execute
act=ActionChains(driver)
act.drag_and_drop_by_offset(min,100,0).perform()  # 100+59 - x value
act.drag_and_drop_by_offset(max,-39,0).perform()   #421-39  - x value

print(min.location)
print(max.location)

driver.quit()

"""
# Scrolling operation
"""
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://www.countries-ofthe-world.com/flags-of-the-world.html")
driver.maximize_window()



# 1. scroll down page by pixel

driver.execute_script("window.scrollBy(0,3000)","")   # 3000 is value
value=driver.execute_script("return window.pageYOffset;")
print("N.o of pixel moved",value)
time.sleep(5)
driver.quit()


#2. Scroll down page till the element is visible

flag=driver.find_element(By.XPATH,"//img[@alt='Flag of India']")
driver.execute_script("arguments[0].scrollIntoView();", flag)
value=driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:",value)
time.sleep(5)
driver.quit()

# 3. SCroll down page till end

driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")
value=driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:",value)


#4.Scrool to starting position

time.sleep(5)  # Wait for 5 seconds
driver.execute_script("window.scrollBy(0,-document.body.scrollHeight)")
value = driver.execute_script("return window.pageYOffset;")
print("Number of pixels moved:", value)

driver.quit()

"""

"""
#### DAY -12 #####

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://text-compare.com/")

# KEYBOARD ACTION

# cltr+a ctrl+c tab ctrl+v

input1 = driver.find_element(By.ID, "inputText1")
input2 = driver.find_element(By.ID, "inputText2")

name = "you are welcome"
input1.send_keys(name)

act = ActionChains(driver)

act.key_down(Keys.CONTROL)  # press ctrl key
act.send_keys("a")  # press a
act.key_up(Keys.CONTROL)  # [release the key]
act.perform()

# Write in a single line

act.key_down(Keys.CONTROL).send_keys("c").key_up(Keys.CONTROL).perform()

# press tab key to navigate to input

act.send_keys(Keys.TAB).perform()

# paste

act.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()

driver.find_element(By.XPATH, "//div[@class='compareButtonText']").click()

print("Entered Text:", name)

time.sleep(10)  # Adding a delay for demonstration purposes
driver.quit()

"""

# HOW TO DOWNLOAD THE FILE

# FIREFOX BROWSER

"""SKIPPED THE TOPIC - FILE DOWNLOAD """


"""

# UPLOAD THE FILE SCENARIO

import time
from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://www.foundit.in/")
driver.find_element(By.XPATH,"buttonContainer_secondaryBtn__text").click()
driver.find_element(By.XPATH,"//input[@id='file-upload']").send_keys("C:\\Users\\a250580\\OneDrive - Syneos Health\\Pictures\\sample.pdf")

driver.quit()


#PIM - ADD EMPLOYEE


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


driver = webdriver.Chrome()

driver.get("https://opensource-demo.orangehrmlive.com/")
driver.implicitly_wait(10)

driver.find_element(By.NAME, "username").send_keys("Admin")
driver.find_element(By.NAME,"password").send_keys("admin123")
driver.find_element(By.XPATH, "//button[@type='submit']").click()
driver.find_element(By.XPATH,"//span[normalize-space()='PIM']").click()
driver.find_element(By.XPATH, "//a[normalize-space()='Add Employee']").click()

# Click on the button to open the file dialog
overlay = WebDriverWait(driver, 2).until(EC.invisibility_of_element_located((By.CLASS_NAME, "oxd-form-loader")))

# Click on the button to open the file dialog
file_input_button = driver.find_element(By.XPATH, "//i[@class='oxd-icon bi-plus']")
file_input_button.click()

# Locate the file input element inside the file dialog
file_input = driver.find_element(By.XPATH, "//input[@type='file']")

# Send the file path to the file input element
file_path = "C://Users//a250580//OneDrive - Syneos Health//Pictures//Screenshots//https status code.png"
file_input.send_keys(file_path)# Wait for the file to be uploaded

time.sleep(25)
driver.quit()




# DAY-13 #####
# How to capture the screenshot of the web page
     1. get_screenshot_as_file
     2. save_screenshot
     3. get_screenshot_as_png
     4. get_screenshot_as_png    # save snap as binary format

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
driver.get('https://www.dummyticket.com/dummy-ticket-for-visa-application/')
driver.find_element(By.XPATH,"//span[@id='select2-billing_country-container']").click()
countries=driver.find_elements(By.XPATH,"//ul[@id='select2-billing_country-results']/li")
print(len(countries))
driver.implicitly_wait(5)
for coun in countries:
    if coun.text=="India":
        coun.click()
        break
time.sleep(20)
driver.quit()



## BOOSTRAP DROPDOWN

import time
from selenium import webdriver
from selenium.webdriver.common.by import By

try:
    driver = webdriver.Chrome()
    driver.implicitly_wait(20)
    driver.get("https://demo.nopcommerce.com/")
    driver.implicitly_wait(20)
    driver.maximize_window()

    # Add a delay to make sure the page is fully loaded before taking a screenshot
    time.sleep(5)

    # How to capture the screenshot of the web page
    driver.save_screenshot("C:\\Users\\a250580\\OneDrive - Syneos Health\\Pictures\\Screenshots\\snap.png")
    print("Screenshot saved successfully!")

except Exception as e:
    print(f"An error occurred: {str(e)}")

finally:
    driver.quit()
    

# How to capture the screenshot of the web page

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import os   # operating sysytem


driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://demo.nopcommerce.com/")
driver.implicitly_wait(20)
driver.maximize_window()

#How to capture the screenshot of the web page

#driver.save_screenshot(os.getcwd()+"\\snap.png")  # get current working directory

#driver.get_screenshot_as_file("C:\\Users\\a250580\\OneDrive - Syneos Health\\Pictures\\Screenshots\\snap.png")

driver.get_screenshot_as_png("C:\\Users\\a250580\\OneDrive - Syneos Health\\Pictures\\Screenshots\\snap.png")  # get snap as binary format

time.sleep(10)
driver.quit()




import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import os   # operating sysytem
from selenium.webdriver import Keys

driver = webdriver.Chrome()
driver.get("https://demo.nopcommerce.com/")
driver.implicitly_wait(20)
driver.maximize_window()

reg=Keys.CONTROL+Keys.RETURN

# Register to new tab

#driver.find_element(By.XPATH,"//a[normalize-space()='Register']").send_keys(reg)




import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import os   # operating sysytem
from selenium.webdriver import Keys

driver = webdriver.Chrome()

# opens a new tab and switched to the new tab

driver.get("https://www.opencart.com")
driver.switch_to.new_window('tab')
driver.get("https://www.orangehrm.com")

# switched to new window

driver.get("https://www.opencart.com")
driver.switch_to.new_window('window')
driver.get("https://www.orangehrm.com")

time.sleep(10)
driver.quit()



# HOW TO HANDLE COOKIES
1. It not a web element
2. single cookies contain multiple attribute
3. Dict format


# COOKIES

1. get_cookies()     - Get all the cookies
2. add_cookies       - Add the cookies
3. delete_cookies()  - Delete the specific ookies
4. delete_all_cookies - Delete all the cookies


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import os   # operating sysytem


driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://demo.nopcommerce.com/")

# CAPTURE COOKIES FROM THE BROWSERS

cookies=driver.get_cookies()
print("size of cookies:",len(cookies))

#for c in cookies:
    #print(c.get('name'),":",c.get('value'))

# ADD NEW COOKIES

driver.add_cookie({"name":"added cookie","value":"123456"})
cookies=driver.get_cookies()
print("size added new of cookies:",len(cookies))

# DELETE SPECIFIC COOKIE FROM BROWSER

driver.delete_cookie("added cookie")
cookies=driver.get_cookies()
print("size delete the cookies:",len(cookies))


# DELETE ALL THE COOKIES

driver.delete_all_cookies()
cookies=driver.get_cookies()
print("size of all deleted cookies:",len(cookies))

driver.quit()

# HEADLESS MODE
Dont see any application in front end but the script run in backend

Advantage
1. Do multiple task
2. Excecution will be fast and reduce time

Disadvantage

1. Do know the application flow


# Different from browser to browser



from selenium import webdriver


def headless_chrome():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')  # Run Chrome in headless mode (without a graphical user interface)

    driver = webdriver.Chrome(options=options)
    return driver


driver = headless_chrome()
driver.get("https://demo.nopcommerce.com/")
print("Title:", driver.title)
print("URL:", driver.current_url)
driver.quit()

#EDGE 


from selenium import webdriver
def headless_edge():
    options=webdriver.ChromeOptions()
    options.add_argument('--headless')

    driver=webdriver.Edge(options=options)
    return driver

driver = headless_edge()
driver.get("https://demo.nopcommerce.com/")
print("Title:", driver.title)
print("URL:", driver.current_url)
driver.quit()

####### DAY - 14 ########

# DATA DRIVEN TESTING

Same test case run every time different sets of data

openpyxl - we can work with excel file (.xlsx)  it is installed by using the command pip install openpyxl
1. Read data from excel
2. how to write data into excel
3. Data driven test case



import openpyxl

# File - workbook -  sheets - Rows - cells

file="C://Users//a250580//OneDrive - Syneos Health//Documents//samplefile.xlsx"
workbook=openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row
column=sheet.max_column

# Read all the row and column

for r in range(1,rows+1):  # outer
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end=",")
    print()


import openpyxl

file="C://Users//a250580//OneDrive - Syneos Health//Documents//sample.xlsx"
workbook=openpyxl.load_workbook(file)
#sheet = workbook["Data"]  # Data also perfectly okay but we have data in all sheet then specify the sheet number
sheet=workbook.active  # if we have only one single sheet at that time we use

#reading
#data=sheet.cell(r,c).value
# writing
#sheet.cell(r,c).value="welcome"

for r in range(1,4):
    for c in range(1,5):
        sheet.cell(r, c).value = "name"
workbook.save(file)



# Multiple data
import openpyxl

file = "C://Users//a250580//OneDrive - Syneos Health//Documents//sample.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1, 1).value = "s.no"
sheet.cell(1, 2).value = "name"
sheet.cell(1, 3).value = "age"

sheet.cell(2, 1).value = 1
sheet.cell(2, 2).value = "pavi"
sheet.cell(2, 3).value = "23"

sheet.cell(3, 1).value = 2
sheet.cell(3, 2).value = "praveen"
sheet.cell(3, 3).value = "25"

workbook.save(file)




import openpyxl
from openpyxl.styles import PatternFill
file="C://Users//a250580//OneDrive - Syneos Health//Documents//sample.xlsx"


import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from Selenium import xl_utilities as xl
from selenium.webdriver.support.ui import Select

driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html")
file="C://Users//a250580//OneDrive - Syneos Health//Documents//sample.xlsx"

rows=xl.getRowCount(file,"Sheet1")
for r in range(2,rows+1):  # 2 3 4 5
   principal=xl.readData(file,"Sheet1",r,1)
   roi=xl.readData(file,"Sheet1",r,2)
   p1=xl.readData(file,"Sheet1",r,3)
   p2=xl.readData(file, "Sheet1", r, 4)
   fre=xl.readData(file,"Sheet1",r,5)
   exp_mau=xl.readData(file,"Sheet1",r,6)

  # Passing data to the application
   driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(principal)
   driver.find_element(By.XPATH,"//input[@id='interest']").send_keys(roi)
   driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(p1)
   prd_ten=Select(driver.find_element(By.XPATH,"//select[@id='tenurePeriod']")) # passing select class
   prd_ten.select_by_visible_text(p2)
   drp_ele=Select(driver.find_element(By.XPATH,"//select[@id='frequency']"))
   drp_ele.select_by_visible_text(fre)
   driver.find_element(By.XPATH,"//img[@src='https://images.moneycontrol.com/images/mf_revamp/btn_calcutate.gif']").click()   # calculate button
   act_mau=driver.find_element(By.XPATH,"//span[@id='resp_matval']/strong").text


   if float(exp_mau)==float(act_mau):
      print("test passed")
      xl.writeData(file,"Sheet1",r,8,"Pass")
      xl.fillGreenColor(file,"Sheet1",r,8)
   else:
      print("test failed")
      xl.writeData(file, "Sheet1", r, 8, "Fail")
      xl.fillRedColor(file, "Sheet1", r, 8)

   driver.find_element(By.XPATH,"//img[@class='PL5']").click()
   time.sleep(2)

# XL Utility file
# Reduce the complexcity
# Reusuability

time.sleep(10)
driver.quit()


# Db server - is a place where the data are stored
# DB Client - light weight software .


SQL 

DDL   - Data Definition lan          -  Create,alter,drop,truncate
DML   - Data Manipulation            - insert , update , delete
DRL   - Data retrivel lan            - select
TCL   - Transcation control          - commit , rollback
DCL   - Data control                 - grant , revoke

Selenium is meant for testing web app but we can able to interact with the db


update_query="update mobile set model='po12' where name='poco'"
delete_query="delete from mobile where rank1=4"
import mysql.connector

con = mysql.connector.connect(host="localhost", port=3306, user="root",password="Chem@9489", database="test1")
curs = con.cursor()  # Create cursor
#curs.execute(update_query)
curs.execute(delete_query) # Execute
con.commit()
con.close()  # Corrected line to close the connection
print("Finished")

# Connection is not established

delete_query="delete from mobile where rank1=4"

try:
    con = mysql.connector.connect(host="localhost", port=3307, user="root", password="Chem@9489", database="test1") #change port n.o for check
    curs = con.cursor()  # Create cursor
    # curs.execute(update_query)
    curs.execute(delete_query)  # Execute
    con.commit()
    con.close()  # Corrected line to close the connection
except:
    print("Connection unsuccessful")

print("Finished")


import mysql.connector
con = mysql.connector.connect(host="localhost", port=3306, user="root", password="Chem@9489", database="test1") #change port n.o for check
curs = con.cursor()  # Create cursor
curs.execute("select * from mobile")  # Execute
for row in curs:
    print(row[0],row[1],row[2],row[3],row[4])
con.commit()
con.close()
"""
# only print the cursor value
import mysql.connector
con = mysql.connector.connect(host="localhost", port=3306, user="root", password="Chem@9489", database="test1") #change port n.o for check
curs = con.cursor()  # Create cursor
curs.execute("select * from mobile")  # Execute
print(curs)
con.commit()
con.close()